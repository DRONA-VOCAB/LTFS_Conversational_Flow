"""
Transcript retrieval utilities for context engineering.

This module loads call transcripts from the Excel file at the project root
(`LTFS Survey calls - 10 transcripts for each categories.xlsx`) and exposes
simple retrieval helpers that return a few relevant examples for the current
turn.

Design goals:
- Be lightweight and robust (no hard dependency on specific Excel schema).
- Work even if sheets/columns change (we just treat each row as a text blob).
- Provide deterministic, fast in‑memory retrieval suitable for each LLM call.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
from openpyxl import load_workbook


logger = logging.getLogger(__name__)


EXCEL_FILENAME = "LTFS Survey calls - 10 transcripts for each categories.xlsx"


@dataclass
class TranscriptExample:
    """Single transcript snippet used as in‑context example."""

    id: str
    sheet: str
    row_index: int
    text: str
    metadata: Dict[str, Any]


_EXAMPLES: Optional[List[TranscriptExample]] = None


def _project_root() -> Path:
    """Return the project root (directory that contains the Excel file)."""
    # backend/app/services/transcript_retrieval.py -> backend/app -> backend -> project root
    return Path(__file__).resolve().parents[3]


def _excel_path() -> Path:
    return _project_root() / EXCEL_FILENAME


def _load_examples() -> List[TranscriptExample]:
    """
    Load all transcript rows from the Excel file into memory.

    We do NOT assume a particular schema. For each non‑empty row we:
    - Concatenate all non‑empty cell values into a single text string.
    - Store sheet name and row index as metadata.
    """
    global _EXAMPLES

    if _EXAMPLES is not None:
        return _EXAMPLES

    path = _excel_path()
    if not path.exists():
        logger.warning(
            "Transcript Excel file not found at %s. Context examples will be empty.",
            path,
        )
        _EXAMPLES = []
        return _EXAMPLES

    logger.info("Loading transcript examples from %s", path)

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    examples: List[TranscriptExample] = []

    for sheet in wb.worksheets:
        sheet_name = sheet.title
        # Heuristic: skip the header row (assumed first row)
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if idx == 1:
                continue

            # Collect non‑empty cell values as strings
            cells = [str(c).strip() for c in row if c not in (None, "")]
            if not cells:
                continue

            text = " ".join(cells)
            # Very short rows are unlikely to be helpful as examples
            if len(text) < 20:
                continue

            ex = TranscriptExample(
                id=f"{sheet_name}-{idx}",
                sheet=sheet_name,
                row_index=idx,
                text=text,
                metadata={
                    "sheet": sheet_name,
                },
            )
            examples.append(ex)

    _EXAMPLES = examples
    logger.info("Loaded %d transcript examples", len(_EXAMPLES))
    return _EXAMPLES


def _tokenize(text: str) -> List[str]:
    """Very simple whitespace tokenizer in lowercase."""
    return text.lower().split()


def _similarity(query: str, doc: str) -> float:
    """
    Naive similarity: Jaccard overlap over token sets.
    This is cheap and avoids bringing in heavy embedding dependencies.
    """
    q_tokens = set(_tokenize(query))
    d_tokens = set(_tokenize(doc))
    if not q_tokens or not d_tokens:
        return 0.0
    intersection = len(q_tokens & d_tokens)
    union = len(q_tokens | d_tokens)
    return intersection / union if union else 0.0


def get_examples_for_turn(
    session: Dict[str, Any],
    user_input: str,
    phase: str,
    k: int = 3,
) -> List[TranscriptExample]:
    """
    Return up to k transcript examples that are roughly relevant to this turn.

    Current strategy:
    - Load all examples from Excel (once per process).
    - Compute a naive token‑overlap similarity with the user input.
    - Slightly bias towards examples whose sheet name loosely matches the phase.
    """
    examples = _load_examples()
    if not examples or not user_input:
        return []

    phase_lower = (phase or "").lower()

    scored: List[Tuple[float, TranscriptExample]] = []
    for ex in examples:
        base_score = _similarity(user_input, ex.text)

        # Very light heuristic: favor sheets whose name appears in the phase
        sheet_lower = ex.sheet.lower()
        if phase_lower and sheet_lower in phase_lower:
            base_score += 0.05

        scored.append((base_score, ex))

    scored.sort(key=lambda pair: pair[0], reverse=True)
    # Filter out completely zero‑score examples; if all are zero, just take first k
    non_zero = [ex for score, ex in scored if score > 0]
    selected = (non_zero or [ex for _, ex in scored])[:k]

    return selected


__all__ = [
    "TranscriptExample",
    "get_examples_for_turn",
]

